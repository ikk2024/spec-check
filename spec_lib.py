"""规范库数据加载与对比逻辑 - 索引优化版"""

import os
import re
from datetime import datetime
from openpyxl import load_workbook


class SpecLib:
    """规范库管理类，负责加载Excel数据、解析输入、对比匹配"""

    def __init__(self, excel_path=None):
        self.records = []
        self.excel_path = excel_path
        self.filename = ""
        self.record_count = 0
        self.update_time = ""
        # 索引字典，加速查找
        self._code_index = {}   # 标准化编号 -> record
        self._name_index = {}   # 标准化名称 -> record
        if excel_path and os.path.exists(excel_path):
            self.load(excel_path)

    def load(self, excel_path):
        """加载Excel文件，解析规范库数据并构建索引"""
        self.excel_path = excel_path
        self.filename = os.path.basename(excel_path)
        self.records = []
        self._code_index = {}
        self._name_index = {}

        # 先用普通模式读取（需要检测字体格式如删除线）
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            # 提取值
            category = row[0].value if len(row) > 0 else None
            name = row[4].value if len(row) > 4 else None
            code = row[5].value if len(row) > 5 else None
            status = row[6].value if len(row) > 6 else None
            impl_date = row[8].value if len(row) > 8 else None

            # 检测删除线：名称或编号单元格有删除线则标记
            name_strike = False
            code_strike = False
            if len(row) > 4 and row[4].font and row[4].font.strike:
                name_strike = True
            if len(row) > 5 and row[5].font and row[5].font.strike:
                code_strike = True

            if name or code:
                record = {
                    "category": str(category).strip() if category else "",
                    "name": str(name).strip() if name else "",
                    "code": str(code).strip() if code else "",
                    "status": str(status).strip() if status else "",
                    "impl_date": self._format_date(impl_date),
                    "strike": name_strike or code_strike,
                }
                self.records.append(record)

                # 构建索引（现行记录优先，其次有误/废止，最后无status）
                norm_code = self._normalize_code(record["code"])
                if norm_code:
                    existing = self._code_index.get(norm_code)
                    if not existing:
                        self._code_index[norm_code] = record
                    elif record["status"] == "现行" and existing["status"] != "现行":
                        self._code_index[norm_code] = record
                    elif not existing["status"] and record["status"]:
                        self._code_index[norm_code] = record
                norm_name = self._normalize_name(record["name"])
                if norm_name:
                    existing = self._name_index.get(norm_name)
                    if not existing:
                        self._name_index[norm_name] = record
                    elif record["status"] == "现行" and existing["status"] != "现行":
                        self._name_index[norm_name] = record
                    elif not existing["status"] and record["status"]:
                        self._name_index[norm_name] = record

        wb.close()
        self.record_count = len(self.records)
        self.update_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    @staticmethod
    def _format_date(date_val):
        if date_val is None:
            return ""
        if isinstance(date_val, datetime):
            return date_val.strftime("%Y-%m-%d")
        return str(date_val)

    @staticmethod
    def _normalize_code(code):
        if not code:
            return ""
        s = code.strip().replace(" ", "").replace("—", "-").replace("–", "-").replace("−", "-")
        return s

    @staticmethod
    def _normalize_name(name):
        if not name:
            return ""
        return name.strip().replace(" ", "").replace("\u3000", "")

    def parse_input(self, text):
        """解析多行输入文本，提取每行的名称和编号"""
        results = []
        lines = text.strip().split("\n")
        for line in lines:
            line = line.strip()
            if not line:
                continue
            results.append(self._parse_line(line))
        return results

    def _parse_line(self, line):
        """解析单行文本，提取编号和名称"""
        # 预处理：将tab替换为空格，合并多余空格
        line = line.replace("\t", " ").strip()
        line = re.sub(r"\s+", " ", line)
        # 去除书名号《》，只保留中间的名称
        line = re.sub(r"[《〈]", "", line)
        line = re.sub(r"[》〉]", "", line)

        # 按优先级排列的编号匹配模式
        patterns = [
            # 地方标准/文件号：闽建科[2005]50号、建标[2015] 38号、闽消【2024】109号
            r"[\u4e00-\u9fff]{1,6}[\[【]\d{4}[\]】]\s*\d+号?",
            # 图集编号：07CJ03-1、22G614-1、10SG614-2、07FJ02
            r"\d{2}[A-Z]{1,3}\d{2,3}[\-—–]?\d*",
            # 国标/行标：GB55001-2021、JGJ/T 229-2010、CJJ 83-2016
            r"[A-Z]{1,4}/?T?\s*\d{1,5}[\-—–]\d{2,4}",
            # DBJ地方标准：DBJT13-118、DBJ/T13-233
            r"DBJ/?T?\s*\d{1,3}[\-—–]\d{1,4}",
            # 简写编号：GB/T 50326
            r"[A-Z]{1,4}/?T?\s*\d{3,6}",
        ]

        code = ""
        name = ""

        for pattern in patterns:
            match = re.search(pattern, line)
            if match:
                code = match.group(0).strip()
                name = line[:match.start()] + line[match.end():]
                name = re.sub(r"^[\s,，、：:]+", "", name)
                name = re.sub(r"[\s,，、：:]+$", "", name)
                break

        if not code:
            # 尝试匹配尾部 -年份 格式（如"福建省绿色建筑设计标准 -2022"）
            tail_match = re.search(r"\s+[\-—–]\s*(\d{4})\s*$", line)
            if tail_match:
                # 保留完整输入名称，年份部分也作为名称的一部分
                name = line.strip()
                code = ""
            else:
                name = line

        return {"input_name": name, "input_code": code}

    def _find_current_version(self, matched):
        """查找废止规范的现行替代版本"""
        # 兼容两种字典格式：_compare_single返回的match_name/match_code，或record的name/code
        name = matched.get("match_name", "") or matched.get("name", "")
        code = matched.get("match_code", "") or matched.get("code", "")
        if not name and not code:
            return None

        norm_name = self._normalize_name(name)
        norm_code = self._normalize_code(code)

        # 从编号中提取核心部分（去掉年份和版本号），如10J113-1 → J113-1
        code_core = re.sub(r'^\d{2}', '', norm_code) if norm_code else ""

        # 在所有记录中查找：名称相似 + 状态为现行
        best_match = None
        best_score = 0
        for rec in self.records:
            if rec["status"] != "现行":
                continue
            rec_norm_name = self._normalize_name(rec["name"])
            rec_norm_code = self._normalize_code(rec["code"])

            # 编号核心部分匹配（如J113-1）
            rec_code_core = re.sub(r'^\d{2}', '', rec_norm_code) if rec_norm_code else ""
            if code_core and rec_code_core and code_core == rec_code_core:
                # 编号核心完全匹配，优先级最高
                return rec

            # 名称匹配
            if norm_name and rec_norm_name:
                if norm_name == rec_norm_name:
                    score = 100
                elif norm_name in rec_norm_name or rec_norm_name in norm_name:
                    score = 80
                else:
                    continue
                if score > best_score:
                    best_score = score
                    best_match = rec

        return best_match

    def compare(self, items):
        """对比已解析的条目列表与规范库

        参数 items: list of {"input_name": str, "input_code": str}
        当匹配到废止/有误规范时，自动在其上方插入现行替代版本
        """
        results = []
        for item in items:
            result = self._compare_single(item.get("input_name", ""), item.get("input_code", ""))
            result["input_name"] = item.get("input_name", "")
            result["input_code"] = item.get("input_code", "")

            # 如果废止或有误，查找现行替代版本并插入
            if result["status"] in ("废止", "有误") and result["match_name"]:
                current = self._find_current_version(result)
                if current:
                    current_row = {
                        "input_name": result["input_name"],
                        "input_code": result["input_code"],
                        "match_name": current["name"],
                        "match_code": current["code"],
                        "name_correct": "是",
                        "code_correct": "否",
                        "status": "现行",
                        "impl_date": current["impl_date"],
                        "is_replacement": True,
                    }
                    results.append(current_row)

            results.append(result)
        return results

    def _find_current_by_name(self, name):
        """根据名称查找现行版本"""
        if not name:
            return None
        norm_name = self._normalize_name(name)
        for rec in self.records:
            if rec["status"] == "现行" and self._normalize_name(rec["name"]) == norm_name:
                return rec
        return None

    def _fuzzy_match_code(self, norm_code):
        """编号模糊匹配：找最相似的记录，优先现行版本"""
        if not norm_code or len(norm_code) < 3:
            return None

        candidates = []
        for key, rec in self._code_index.items():
            if not key or len(key) < 3:
                continue
            # 输入编号包含库中编号，或库中编号包含输入编号
            if key in norm_code or norm_code in key:
                # 计算相似度：重合长度占比
                overlap = min(len(key), len(norm_code))
                score = overlap / max(len(key), len(norm_code))
                # 现行版本加分
                if rec["status"] == "现行":
                    score += 0.1
                candidates.append((score, rec))

        if not candidates:
            return None
        # 按相似度降序排列，取最高分
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    def _fuzzy_match_name(self, norm_name):
        """名称模糊匹配：找最相似的记录，优先现行版本"""
        if not norm_name or len(norm_name) < 2:
            return None

        candidates = []
        for key, rec in self._name_index.items():
            if not key or len(key) < 2:
                continue
            # 包含匹配
            if key in norm_name or norm_name in key:
                overlap = min(len(key), len(norm_name))
                score = overlap / max(len(key), len(norm_name))
                if rec["status"] == "现行":
                    score += 0.1
                candidates.append((score, rec))
                continue
            # 关键词交集匹配：累计所有不重叠的公共子串长度
            common_len = 0
            used = set()
            # 按子串长度降序查找，贪心匹配
            for sub_len in range(min(len(norm_name), len(key)), 1, -1):
                for i in range(len(norm_name) - sub_len + 1):
                    sub = norm_name[i:i + sub_len]
                    if sub in key:
                        # 检查是否与已匹配部分重叠
                        overlap = False
                        for s, e in used:
                            if i < e and (i + sub_len) > s:
                                overlap = True
                                break
                        if not overlap:
                            common_len += sub_len
                            used.add((i, i + sub_len))
            if common_len >= 2:
                score = common_len / max(len(key), len(norm_name))
                if rec["status"] == "现行":
                    score += 0.1
                candidates.append((score, rec))

        if not candidates:
            return None
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    def _compare_single(self, input_name, input_code):
        """对比单条记录（使用索引加速）"""
        match_by_code = None
        match_by_name = None

        # 优先按编号索引查找
        if input_code:
            norm_code = self._normalize_code(input_code)
            match_by_code = self._code_index.get(norm_code)
            # 精确匹配不到，做模糊匹配
            if not match_by_code:
                match_by_code = self._fuzzy_match_code(norm_code)

        # 按名称索引查找
        if input_name:
            norm_name = self._normalize_name(input_name)
            match_by_name = self._name_index.get(norm_name)
            # 精确匹配不到，做模糊匹配
            if not match_by_name:
                match_by_name = self._fuzzy_match_name(norm_name)

        # 综合判断：编号和名称都匹配到时，优先取名称匹配（更可靠）
        # 如果编号匹配到的记录名称与输入不相关，则用名称匹配结果
        matched = match_by_code or match_by_name
        if match_by_code and match_by_name:
            # 编号和名称都匹配到了不同记录，优先名称匹配
            if match_by_code != match_by_name:
                # 检查编号匹配结果的名称是否与输入相关
                if input_name:
                    norm_input = self._normalize_name(input_name)
                    norm_code_match = self._normalize_name(match_by_code["name"])
                    if norm_input not in norm_code_match and norm_code_match not in norm_input:
                        matched = match_by_name
                    else:
                        matched = match_by_code
                else:
                    matched = match_by_code
            else:
                matched = match_by_code

        if not matched:
            return {
                "match_name": "",
                "match_code": "",
                "name_correct": "否",
                "code_correct": "否",
                "status": "未收录",
                "impl_date": "",
            }

        # 判断名称正确性
        name_correct = "否"
        if input_name and matched["name"]:
            norm_input = self._normalize_name(input_name)
            norm_match = self._normalize_name(matched["name"])
            if norm_input == norm_match:
                name_correct = "是"
            elif norm_input in norm_match or norm_match in norm_input:
                name_correct = "部分正确"
            else:
                # 检查是否有2字以上公共子串
                common = 0
                for i in range(len(norm_input) - 1):
                    for j in range(i + 2, len(norm_input) + 1):
                        sub = norm_input[i:j]
                        if len(sub) >= 2 and sub in norm_match:
                            common = max(common, len(sub))
                if common >= 2:
                    name_correct = "部分正确"

        # 判断编号正确性
        code_correct = "否"
        if input_code and matched["code"]:
            norm_input = self._normalize_code(input_code)
            norm_match = self._normalize_code(matched["code"])
            if norm_input == norm_match:
                code_correct = "是"

        if not input_name and matched["name"]:
            name_correct = "是"
        if not input_code and matched["code"]:
            code_correct = "是"

        status = matched["status"] if matched["status"] else "未标注"

        return {
            "match_name": matched["name"],
            "match_code": matched["code"],
            "name_correct": name_correct,
            "code_correct": code_correct,
            "status": status,
            "impl_date": matched["impl_date"],
        }

    def get_status(self):
        return {
            "filename": self.filename,
            "record_count": self.record_count,
            "update_time": self.update_time,
        }
